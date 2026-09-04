#!/usr/bin/env python3
"""生データからシミュレータ用のデータセットを生成する.

入力 (data/raw/):
  jma_stations.json  気象庁 震度観測点 (緯度経度)
  jmacode.zip        気象庁 電文コード表 (震央地名・細分区域・津波予報区)
  avs30_cache.json   J-SHIS 表層地盤 (AVS30 / 増幅率 / 微地形区分)
  japan.geojson      都道府県界ポリゴン

出力 (web/data/):
  stations.json      観測点 (座標・地盤・所属震央地名) の列指向データ
  regions.json       震央地名 (コード・名称・代表座標・陸海種別)
  tsunami_zone_list.json  津波予報区の一覧 (沿岸線は build_tsunami_zones.py で付与)
  japan.geojson      簡略化した都道府県界
  traveltime.json    P/S 走時表 (深さ x 震央距離)
"""

from __future__ import annotations

import argparse
import io
import json
import sys
import zipfile
from pathlib import Path

import numpy as np

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from sim.velocity import VelocityModel, travel_time

ROOT = Path(__file__).resolve().parent.parent
RAW = ROOT / "data" / "raw"
OUT = ROOT / "web" / "data"

CODE_XLSX = "地震火山関連コード表.xlsx"

# 細分区域名 -> 震央地名 の個別対応 (語尾に「地方」を付けるだけでは一致しないもの)
SUBAREA_OVERRIDES = {
    "三宅島": "三宅島近海",
    "伊豆大島": "伊豆大島近海",
    "八丈島": "八丈島近海",
    "新島": "新島・神津島近海",
    "神津島": "新島・神津島近海",
    "兵庫県淡路島": "淡路島付近",
    "北海道利尻礼文": "北海道北西沖",
    "北海道奥尻島": "北海道南西沖",
    "島根県隠岐": "隠岐島近海",
    "新潟県佐渡": "佐渡付近",
    "長崎県五島": "五島列島近海",
    "長崎県壱岐": "壱岐・対馬近海",
    "長崎県対馬": "壱岐・対馬近海",
    "沖縄県与那国島": "与那国島近海",
    "沖縄県久米島": "沖縄本島近海",
    "沖縄県大東島": "南大東島近海",
    "沖縄県宮古島": "宮古島近海",
    "沖縄県本島中南部": "沖縄本島近海",
    "沖縄県本島北部": "沖縄本島近海",
    "沖縄県石垣島": "石垣島近海",
    "沖縄県西表島": "西表島付近",
    "鹿児島県十島村": "トカラ列島近海",
    "鹿児島県奄美北部": "奄美大島近海",
    "鹿児島県奄美南部": "奄美大島近海",
    "鹿児島県屋久島": "種子島近海",
    "鹿児島県甑島": "薩摩半島西方沖",
    "鹿児島県種子島": "種子島近海",
}

DEFAULT_AVS30 = 350.0  # J-SHIS メッシュ外 (離島など) の既定値

# 名称の末尾がこれらの震央地名は海域区分として扱う。
# 離島の「〜近海」などは島に観測点があるため陸域の重心が求まるが、
# 沖合の地点からも引けるように海域アンカーにも登録する。
SEA_SUFFIXES = ("沖", "湾", "灘", "海峡", "近海", "付近", "水道", "海")


# ---------------------------------------------------------------- コード表
def load_code_tables() -> dict:
    import openpyxl

    with zipfile.ZipFile(RAW / "jmacode.zip") as z:
        wb = openpyxl.load_workbook(io.BytesIO(z.read(CODE_XLSX)), read_only=True)

    epicenter = {}
    for r in list(wb["41"].iter_rows(values_only=True))[3:]:
        if r[0] and r[1]:
            epicenter[str(r[0])] = str(r[1])

    # 細分区域 -> 観測点名 の対応
    station_area: dict[str, str] = {}
    subareas: set[str] = set()
    for r in list(wb["24"].iter_rows(values_only=True))[3:]:
        if r[1] and r[7]:
            subareas.add(str(r[1]))
            station_area[str(r[7])] = str(r[1])

    tsunami = []
    for r in list(wb["31"].iter_rows(values_only=True))[3:]:
        if r[0] and r[1]:
            kind = str(r[3]) if len(r) > 3 and r[3] else ""
            tsunami.append(
                {"code": str(r[0]), "name": str(r[1]), "kana": str(r[2] or ""), "kind": kind}
            )

    return {
        "epicenter": epicenter,
        "station_area": station_area,
        "subareas": subareas,
        "tsunami": tsunami,
    }


def map_subarea_to_epicenter(subareas: set[str], epicenter_names: set[str]) -> dict[str, str]:
    """細分区域名を震央地名に対応づける。"""
    out = {}
    unresolved = []
    for name in sorted(subareas):
        if name in epicenter_names:
            out[name] = name
        elif name in SUBAREA_OVERRIDES:
            out[name] = SUBAREA_OVERRIDES[name]
        elif name + "地方" in epicenter_names:
            out[name] = name + "地方"
        else:
            unresolved.append(name)
    if unresolved:
        print(f"  [警告] 震央地名に対応づかない細分区域: {unresolved}")
    return out


# ---------------------------------------------------------------- 観測点
def build_stations(codes: dict) -> dict:
    stations = json.loads((RAW / "jma_stations.json").read_text(encoding="utf-8"))
    avs_cache = json.loads((RAW / "avs30_cache.json").read_text(encoding="utf-8"))

    epi_names = set(codes["epicenter"].values())
    sub2epi = map_subarea_to_epicenter(codes["subareas"], epi_names)
    name2code = {v: k for k, v in codes["epicenter"].items()}

    lat, lon, avs, arv, region, names, pref, geomorph = [], [], [], [], [], [], [], []
    missing_avs = 0
    missing_region = 0

    for st in stations:
        la, lo = float(st["lat"]), float(st["lon"])
        key = f"{la:.4f},{lo:.4f}"
        site = avs_cache.get(key)
        if site is None:
            missing_avs += 1
            a30, a_rv, gm = DEFAULT_AVS30, 0.0, ""
        else:
            a30, a_rv, gm = site["avs30"], site["arv"], site.get("geomorph", "")

        sub = codes["station_area"].get(st["name"], "")
        epi_name = sub2epi.get(sub, "")
        code = name2code.get(epi_name, "")
        if not code:
            missing_region += 1

        lat.append(round(la, 4))
        lon.append(round(lo, 4))
        avs.append(round(float(a30), 1))
        arv.append(round(float(a_rv), 2))
        region.append(code)
        names.append(st["name"])
        pref.append(int(st.get("pref", 0) or 0))
        geomorph.append(gm)

    print(
        f"  観測点 {len(stations)} 点 / AVS30 欠測 {missing_avs} 点 "
        f"/ 震央地名未割当 {missing_region} 点"
    )
    return {
        "count": len(stations),
        "lat": lat,
        "lon": lon,
        "avs30": avs,
        "arv": arv,
        "region": region,
        "name": names,
        "pref": pref,
        "geomorph": geomorph,
    }


# ---------------------------------------------------------------- 震央地名
def build_regions(codes: dict, stations: dict) -> dict:
    sea = json.loads((ROOT / "data" / "sea_region_coords.json").read_text(encoding="utf-8"))
    sea = {k: v for k, v in sea.items() if not k.startswith("_")}

    # 陸域の震央地名は所属観測点の重心を代表座標とする
    acc: dict[str, list[list[float]]] = {}
    for la, lo, code in zip(stations["lat"], stations["lon"], stations["region"]):
        if code:
            acc.setdefault(code, []).append([la, lo])

    regions = []
    for code, name in sorted(codes["epicenter"].items()):
        pts = acc.get(code)
        anchors: list[list[float]] = []
        if pts:
            arr = np.array(pts)
            lat_c, lon_c = float(arr[:, 0].mean()), float(arr[:, 1].mean())
            # 離島の海域区分は、観測点の重心を海域アンカーとしても登録する
            if name.endswith(SEA_SUFFIXES):
                kind = "sea"
                anchors = [[round(lat_c, 4), round(lon_c, 4)]]
                if code in sea:
                    extra = sea[code]
                    if not isinstance(extra[0], list):
                        extra = [extra]
                    anchors += [[round(float(a), 4), round(float(b), 4)] for a, b in extra]
            else:
                kind = "land"
        elif code in sea:
            pts_sea = sea[code]
            if not isinstance(pts_sea[0], list):
                pts_sea = [pts_sea]
            arr = np.array(pts_sea, dtype=float)
            lat_c, lon_c = float(arr[:, 0].mean()), float(arr[:, 1].mean())
            anchors = [[round(float(a), 4), round(float(b), 4)] for a, b in arr]
            kind = "sea"
        else:
            continue  # 国外・遠地は代表座標を持たない
        entry = {
            "code": code,
            "name": name,
            "lat": round(lat_c, 4),
            "lon": round(lon_c, 4),
            "type": kind,
            "stations": len(pts) if pts else 0,
        }
        if kind == "sea":
            entry["anchors"] = anchors
        regions.append(entry)

    land = sum(1 for r in regions if r["type"] == "land")
    print(f"  震央地名 {len(regions)} 件 (陸域 {land} / 海域・広域 {len(regions) - land})")
    return {"regions": regions}


# ---------------------------------------------------------------- 地図
def douglas_peucker(points: np.ndarray, tol: float) -> np.ndarray:
    """Douglas-Peucker によるポリライン簡略化。"""
    if points.shape[0] < 3:
        return points
    keep = np.zeros(points.shape[0], dtype=bool)
    keep[0] = keep[-1] = True
    stack = [(0, points.shape[0] - 1)]
    while stack:
        i, j = stack.pop()
        if j <= i + 1:
            continue
        seg = points[j] - points[i]
        norm = np.hypot(*seg)
        rel = points[i + 1 : j] - points[i]
        if norm < 1e-12:
            d = np.hypot(rel[:, 0], rel[:, 1])
        else:
            d = np.abs(seg[0] * rel[:, 1] - seg[1] * rel[:, 0]) / norm
        k = int(np.argmax(d))
        if d[k] > tol:
            k += i + 1
            keep[k] = True
            stack.append((i, k))
            stack.append((k, j))
    return points[keep]


def ring_area(ring: np.ndarray) -> float:
    x, y = ring[:, 0], ring[:, 1]
    return 0.5 * abs(float(np.dot(x, np.roll(y, 1)) - np.dot(y, np.roll(x, 1))))


def simplify_geojson(tol: float, min_area: float, ndigits: int) -> dict:
    src = json.loads((RAW / "japan.geojson").read_text(encoding="utf-8"))
    out_features = []
    n_in = n_out = 0

    for feat in src["features"]:
        geom = feat["geometry"]
        polys = geom["coordinates"] if geom["type"] == "MultiPolygon" else [geom["coordinates"]]
        new_polys = []
        for poly in polys:
            new_rings = []
            for ri, ring in enumerate(poly):
                arr = np.array(ring, dtype=float)
                n_in += arr.shape[0]
                if ri == 0 and ring_area(arr) < min_area:
                    break  # 微小な島は落とす
                simp = douglas_peucker(arr, tol)
                if simp.shape[0] < 4:
                    if ri == 0:
                        break
                    continue
                simp = np.round(simp, ndigits)
                n_out += simp.shape[0]
                new_rings.append(simp.tolist())
            if new_rings:
                new_polys.append(new_rings)
        if not new_polys:
            continue
        out_features.append(
            {
                "type": "Feature",
                "properties": {
                    "name": feat["properties"].get("nam_ja", ""),
                    "id": feat["properties"].get("id"),
                },
                "geometry": {"type": "MultiPolygon", "coordinates": new_polys},
            }
        )

    print(f"  地図: 頂点 {n_in} -> {n_out} ({100 * n_out / max(n_in, 1):.1f}%)")
    return {"type": "FeatureCollection", "features": out_features}


# ---------------------------------------------------------------- 走時表
def build_traveltime_table(
    depths: list[float], max_distance: float = 1200.0, n_dist: int = 241
) -> dict:
    model = VelocityModel()
    dist = np.linspace(0.0, max_distance, n_dist)
    table = {"depths": depths, "distances": dist.tolist(), "p": [], "s": []}
    for d in depths:
        table["p"].append([round(float(v), 3) for v in travel_time(model, d, "P").time(dist)])
        table["s"].append([round(float(v), 3) for v in travel_time(model, d, "S").time(dist)])
    print(f"  走時表: 深さ {len(depths)} 段 x 距離 {n_dist} 点")
    return table


# ---------------------------------------------------------------- main
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--map-tolerance", type=float, default=0.0025, help="地図の簡略化許容誤差 [度]")
    ap.add_argument("--map-min-area", type=float, default=2e-5, help="残す島の最小面積 [平方度]")
    ap.add_argument("--map-digits", type=int, default=4)
    args = ap.parse_args()

    OUT.mkdir(parents=True, exist_ok=True)

    print("コード表を読み込み中...")
    codes = load_code_tables()

    print("観測点データを構築中...")
    stations = build_stations(codes)
    (OUT / "stations.json").write_text(
        json.dumps(stations, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print("震央地名を構築中...")
    regions = build_regions(codes, stations)
    (OUT / "regions.json").write_text(
        json.dumps(regions, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print("津波予報区の一覧を書き出し中...")
    zones = [t for t in codes["tsunami"] if "領域表現（" not in t["kind"]]
    # 沿岸線は tools/build_tsunami_zones.py が付与するため、
    # ここでは中間ファイルとして data/ 側に書き出す
    (ROOT / "data" / "tsunami_zone_list.json").write_text(
        json.dumps({"zones": zones}, ensure_ascii=False, separators=(",", ":")),
        encoding="utf-8",
    )
    print(f"  津波予報区 {len(zones)} 区 (沿岸線は build_tsunami_zones.py で付与)")

    print("地図を簡略化中...")
    gj = simplify_geojson(args.map_tolerance, args.map_min_area, args.map_digits)
    (OUT / "japan.geojson").write_text(
        json.dumps(gj, ensure_ascii=False, separators=(",", ":")), encoding="utf-8"
    )

    print("走時表を作成中...")
    depths = [2, 5, 10, 15, 20, 30, 40, 50, 60, 80, 100, 150, 200, 300, 400, 500, 600]
    tt = build_traveltime_table([float(d) for d in depths])
    (OUT / "traveltime.json").write_text(
        json.dumps(tt, separators=(",", ":")), encoding="utf-8"
    )

    print("\n出力:")
    for f in sorted(OUT.glob("*")):
        print(f"  {f.name:22s} {f.stat().st_size / 1024:8.1f} KB")
    return 0


if __name__ == "__main__":
    sys.exit(main())
